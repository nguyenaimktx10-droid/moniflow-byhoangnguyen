"""One-off: parse Vietcombank-style xlsx, filter Facebook Ads debit rows."""
import json
import sys
from datetime import datetime

import openpyxl

# Default path; override with argv[1]
DEFAULT_XLSX = r"c:\Users\ADMIN\Downloads\Telegram Desktop\SaoKeTK_01032026_30032026.xlsx"


def parse_amount(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(round(v))
    s = str(v).replace(",", "").replace(" ", "").strip()
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def is_facebook_ads(desc: str) -> bool:
    if not desc:
        return False
    d = str(desc)
    u = d.upper()
    if "FACEBK" in u:
        return True
    if "FB.ME/ADS" in u:
        return True
    if "FB.ME" in u and "ADS" in u:
        return True
    return False


def parse_date(b):
    if b is None:
        return None
    if isinstance(b, datetime):
        return b
    s = str(b).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass
    return None


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    rows_out = []
    month_totals: dict[str, int] = {}

    # Data rows after header; Vietcombank template ~ from row 37
    for ri in range(37, ws.max_row + 1):
        b = ws.cell(ri, 2).value  # B date
        r = ws.cell(ri, 18).value  # R details
        af = ws.cell(ri, 32).value  # AF debit
        amt = parse_amount(af)
        if not is_facebook_ads(r) or amt is None or amt <= 0:
            continue
        dt = parse_date(b)
        if not dt:
            continue
        tx = ws.cell(ri, 22).value
        rows_out.append(
            {
                "date": dt.strftime("%Y-%m-%d"),
                "amount": amt,
                "note": ("FB Ads — " + str(r)[:120]) if r else "FB Ads",
                "ref": str(tx) if tx else "",
                "row": ri,
            }
        )
        mk = dt.strftime("%Y-%m")
        month_totals[mk] = month_totals.get(mk, 0) + amt

    out = {
        "source": path,
        "count": len(rows_out),
        "monthTotals": month_totals,
        "grandTotal": sum(month_totals.values()),
        "transactions": rows_out,
    }
    import os

    base = os.path.join(os.path.dirname(__file__), "fb_ads_mar2026.json")
    out_json = base
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(json.dumps({k: out[k] for k in ("count", "monthTotals", "grandTotal")}, ensure_ascii=False))
    print("wrote", out_json)


if __name__ == "__main__":
    main()
